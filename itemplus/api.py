import frappe
import json
from frappe.utils import get_site_path

@frappe.whitelist()
def export_items(items=None, price_list=None, file_format='Excel'):
    """
    Export items to CSV or Excel format where custom_is_weight_item = 1.
    - If no selection (items=None or empty): export all weight items.
    - If selection provided: export only selected items that are weight items.
    - price_list: Price List to use for fetching item prices.
    - file_format: 'Excel' or 'CSV' (default: 'Excel')
    """

    selected_item_names = []

    try:
        # Check if items are passed and non-empty
        if items:
            if isinstance(items, str):
                items = json.loads(items)
            if not isinstance(items, list):
                items = []

            for item in items:
                item_name = item.get("name") if isinstance(item, dict) else item
                if not item_name:
                    continue
                is_weight = frappe.db.get_value("Item", item_name, "custom_is_weight_item")
                if is_weight == 1:
                    selected_item_names.append(item_name)

        # If no items passed or none selected, fallback to all weight items
        if not selected_item_names:
            selected_item_names = frappe.get_all("Item", filters={"custom_is_weight_item": 1}, pluck="name")

        # Still nothing? Then throw an error
        if not selected_item_names:
            return {"error": "No weight items found to export."}

        # Prepare data for both CSV and Excel
        headers = ["Item Code", "Item Name", "Custom Hotkey", "Custom Is Weight Item", "Shelf Life In Days", "Barcode Type", "DB Code", "Item Price"]
        rows = []

        for item_name in selected_item_names:
            item_doc = frappe.get_doc("Item", item_name)
            shelf_life = getattr(item_doc, 'shelf_life_in_days', '') or ''
            barcode_type = '101'
            db_code = '21'
            item_price = ''

            # Fetch item price based on the specified price list
            if price_list:
                filters = {
                    "item_code": item_doc.item_code,
                    "price_list": price_list
                }
                item_price = frappe.db.get_value("Item Price", filters, "price_list_rate") or ''
            else:
                # Fallback to original logic if no price list specified
                cost_center = getattr(item_doc, 'cost_center', None)
                if cost_center:
                    item_price = frappe.db.get_value("Item Price", {
                        "item_code": item_doc.item_code,
                        "selling": 1,
                        "cost_center": cost_center
                    }, "price_list_rate") or ''
                else:
                    item_price = frappe.db.get_value("Item Price", {
                        "item_code": item_doc.item_code,
                        "selling": 1
                    }, "price_list_rate") or ''

            rows.append([
                item_doc.item_code,
                item_doc.item_name,
                getattr(item_doc, 'custom_hotkey', ''),
                getattr(item_doc, 'custom_is_weight_item', ''),
                shelf_life,
                barcode_type,
                db_code,
                item_price
            ])

        # Generate appropriate file format
        if file_format == 'Excel':
            return _generate_excel(headers, rows)
        else:
            return _generate_csv(headers, rows)

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Item Export Failed")
        return {"error": f"Item Export Failed: {str(e)}"}


def _generate_csv(headers, rows):
    """Generate CSV file and return file details."""
    csv_content = "\ufeff"  # UTF-8 BOM for Excel compatibility
    csv_content += ",".join(headers) + "\n"
    
    for row in rows:
        csv_content += ",".join([str(cell) for cell in row]) + "\n"

    # Write file
    file_name = "itemplus.csv"
    file_path = get_site_path("public", "files", file_name)
    with open(file_path, 'w', encoding='utf-8') as f:
        f.write(csv_content)

    return {
        "file_url": f"/files/{file_name}",
        "file_name": file_name
    }


def _generate_excel(headers, rows):
    """Generate Excel file and return file details."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        frappe.throw("openpyxl library is required for Excel export. Please install it.")

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Items"

    # Add headers with formatting
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    # Add data rows
    for row_num, row in enumerate(rows, 2):
        for col_num, value in enumerate(row, 1):
            ws.cell(row=row_num, column=col_num).value = value

    # Adjust column widths
    for col_num, header in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20

    # Write file
    file_name = "itemplus.xlsx"
    file_path = get_site_path("public", "files", file_name)
    wb.save(file_path)

    return {
        "file_url": f"/files/{file_name}",
        "file_name": file_name
    }
